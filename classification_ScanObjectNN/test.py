"""
python test.py --model pointMLP --msg 20220209053148-404
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

import logging
import random
import shutil

import argparse
import os
import datetime

import h5py
import torch
import torch.nn.parallel
import torch.backends.cudnn as cudnn
import torch.optim
import torch.utils.data
import torch.utils.data.distributed
from torch.utils.data import DataLoader
import models as models
from utils import progress_bar, IOStream
from ScanObjectNN import ScanObjectNN
import sklearn.metrics as metrics
from utils import Logger, mkdir_p, progress_bar, save_model, save_args, cal_loss
import numpy as np
import torch.nn.functional as F

model_names = sorted(name for name in models.__dict__
                     if callable(models.__dict__[name]))


def parse_args():
    """Parameters"""
    parser = argparse.ArgumentParser('training')
    parser.add_argument('-c', '--checkpoint', type=str, metavar='PATH',
                        help='path to save checkpoint (default: checkpoint)')
    parser.add_argument('--msg', type=str, default='20231211003033', help='message after checkpoint')
    parser.add_argument('--batch_size', type=int, default=16, help='batch size in training')
    parser.add_argument('--model', default='PointNet', help='model name [default: pointnet_cls]')
    parser.add_argument('--num_classes', default=15, type=int, help='default value for classes of ScanObjectNN')
    # parser.add_argument('--num_classes', default=15, type=int, choices=[10, 40], help='training on ModelNet10/40')
    parser.add_argument('--num_points', type=int, default=1024, help='Point Number')
    parser.add_argument('--seed', default=None, type=int, help='random seed')
    parser.add_argument('-od', '--original_data', default=False, help='use original test dataset')
    return parser.parse_args()

def main():
    args = parse_args()
    if args.seed is None:
        args.seed = np.random.randint(1, 10000)

    # if args.seed is not None:
    #     torch.manual_seed(args.seed)
    #     np.random.seed(args.seed)
    #     torch.cuda.manual_seed_all(args.seed)
    #     torch.cuda.manual_seed(args.seed)
    #     torch.set_printoptions(10)
    #     torch.backends.cudnn.benchmark = False
    #     torch.backends.cudnn.deterministic = True
    #     os.environ['PYTHONHASHSEED'] = str(args.seed)

    time_str = str(datetime.datetime.now().strftime('-%Y%m%d%H%M%S-'))
    args.checkpoint = 'testpoints/' + args.model + time_str + str(args.seed)
    if not os.path.isdir(args.checkpoint):
        mkdir_p(args.checkpoint)

    screen_logger = logging.getLogger()
    screen_logger.setLevel(logging.INFO)
    formatter = logging.Formatter('%(message)s')
    file_handler = logging.FileHandler(os.path.join(args.checkpoint, "out.txt"))
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)
    screen_logger.addHandler(file_handler)

    def printf(str):
        screen_logger.info(str)
        print(str)

    printf(f"args: {args}")
    os.environ["HDF5_USE_FILE_LOCKING"] = "FALSE"

    if torch.cuda.is_available():
        device = 'cuda'
    else:
        device = 'cpu'
    printf(f"==> Using device: {device}")

    # Model
    printf('==> Building model..')
    net = models.__dict__[args.model]()
    criterion = cal_loss
    net = net.to(device)
    checkpoints = 'checkpoints/' + args.model + '-' + args.msg
    checkpoint_path = os.path.join(checkpoints, 'best_checkpoint.pth')
    checkpoint = torch.load(checkpoint_path, map_location=torch.device('cpu'))
    # criterion = criterion.to(device)
    if device == 'cuda':
        net = torch.nn.DataParallel(net)
        cudnn.benchmark = True
    net.load_state_dict(checkpoint['net'])


    printf('==> Preparing data..')

    if args.original_data:
        file_path = 'data/test_objectdataset_augmentedrot_scale75.h5'
        destination_path = 'data/h5_files/main_split'
        try:
            shutil.copy(file_path, destination_path)
        except Exception as e:
            print(f"Failed to move the original file: {e}")
        printf('====original dataset====')
        test_loader = DataLoader(ScanObjectNN(partition='test', num_points=args.num_points), num_workers=4,
                                 batch_size=args.batch_size, shuffle=True, drop_last=False)

        test_out = validate(net, test_loader, criterion, device)
        printf(f"Vanilla out: {test_out}")
    else:
        try:
            # 尝试打开已存在的Excel文件
            workbook = openpyxl.load_workbook('output.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            # 如果文件不存在，创建一个新的Excel文件
            workbook = Workbook()
            sheet = workbook.active
        next_row = sheet.max_row + 1
        sheet[f"A{next_row}"] = args.model + time_str + str(args.seed)
        excel_data = {}
        for noisy in ['gaussian', 'dropout', 'rotation']: # 'gaussian', 'dropout', 'rotation'
            for i in range(1, 5):
                # noisy(i, args.seed)
                if noisy in globals() and callable(globals()[noisy]):
                    func = globals()[noisy]
                    func(i, args.seed)

                printf('====' + noisy + '------' + str(i) + '====')
                test_loader = DataLoader(ScanObjectNN(partition='test', num_points=args.num_points), num_workers=4,
                                         batch_size=args.batch_size, shuffle=True, drop_last=False)

                test_out = validate(net, test_loader, criterion, device)
                printf(f"Vanilla out: {test_out}")
                key_tuple = (noisy, i)
                excel_data[key_tuple] = test_out["acc"]
                # save the noisy dataset
                destination_directory = 'testpoints/' + str(args.seed) + '/' + str(noisy) + '/' + str(i)
                move_files('data/h5_files/main_split', destination_directory)
        write_to_excel(excel_data)
        excel_data.clear()



def validate(net, testloader, criterion, device):
    args = parse_args()
    net.eval()
    test_loss = 0
    correct = 0
    total = 0
    test_true = []
    test_pred = []
    time_cost = datetime.datetime.now()
    with torch.no_grad():
        for batch_idx, (data, label) in enumerate(testloader):
            data, label = data.to(device), label.to(device).squeeze()
            if not args.model == 'PointConT_cls':
                data = data.permute(0, 2, 1)
            # PointCont_cls需要修改

            logits = net(data)
            loss = criterion(logits, label)
            test_loss += loss.item()
            preds = logits.max(dim=1)[1]
            test_true.append(label.cpu().numpy())
            test_pred.append(preds.detach().cpu().numpy())
            total += label.size(0)
            correct += preds.eq(label).sum().item()
            progress_bar(batch_idx, len(testloader), 'Loss: %.3f | Acc: %.3f%% (%d/%d)'
                         % (test_loss / (batch_idx + 1), 100. * correct / total, correct, total))

    time_cost = int((datetime.datetime.now() - time_cost).total_seconds())
    test_true = np.concatenate(test_true)
    test_pred = np.concatenate(test_pred)
    return {
        "loss": float("%.3f" % (test_loss / (batch_idx + 1))),
        "acc": float("%.3f" % (100. * metrics.accuracy_score(test_true, test_pred))),
        "acc_avg": float("%.3f" % (100. * metrics.balanced_accuracy_score(test_true, test_pred))),
        "time": time_cost
    }

def write_to_excel(output_data, excel_file='output.xlsx'):
    try:
        # 尝试打开已存在的Excel文件
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
    except FileNotFoundError:
        # 如果文件不存在，创建一个新的Excel文件
        workbook = Workbook()
        sheet = workbook.active
        # 添加表头
        # headers = ["gaussian-1", "gaussian-2", "gaussian-3", "gaussian-4", "dropout-1", "dropout-2", "dropout-3", "dropout-4", "rotation-1", "rotation-2", "rotation-3", "rotation-4"]
        # for col_num, header in enumerate(headers, 1):
        #     col_letter = get_column_letter(col_num * 2 - 1)  # 插入空列
        #     sheet[f"{col_letter}1"] = header
        #     sheet[f"{col_letter}1"].font = Font(bold=True)

    # 获取下一行的行号
    next_row = sheet.max_row + 1

    # 将输出数据写入Excel文件
    noisy = 'gaussian'
    i = 1
    key_tuple = (noisy, i)
    sheet[f"A{next_row}"] = output_data[key_tuple]
    i += 1
    sheet[f"C{next_row}"] = output_data[key_tuple]
    i += 1
    sheet[f"E{next_row}"] = output_data[key_tuple]
    i += 1
    sheet[f"G{next_row}"] = output_data[key_tuple]

    noisy = 'dropout'
    i = 1
    key_tuple = (noisy, i)
    sheet[f"J{next_row}"] = output_data[key_tuple]
    i += 1
    sheet[f"L{next_row}"] = output_data[key_tuple]
    i += 1
    sheet[f"N{next_row}"] = output_data[key_tuple]
    i += 1
    sheet[f"P{next_row}"] = output_data[key_tuple]

    noisy = 'rotation'
    i = 1
    key_tuple = (noisy, i)
    sheet[f"S{next_row}"] = output_data[key_tuple]
    i += 1
    sheet[f"U{next_row}"] = output_data[key_tuple]
    i += 1
    sheet[f"W{next_row}"] = output_data[key_tuple]
    i += 1
    sheet[f"Y{next_row}"] = output_data[key_tuple]

    # 保存Excel文件
    workbook.save(excel_file)

def move_files(files, destination_directory):
    if not os.path.exists(destination_directory):
        os.makedirs(destination_directory)
    file = ['test_objectdataset_augmentedrot_scale75.h5']
    for file_path in file:
        file_path = os.path.join(files, file_path)
        try:
            shutil.copy(file_path, destination_directory)
        except Exception as e:
            print(f"移动文件 '{file_path}' 失败: {e}")

def gaussian(i, random_seed):
    mean = 0
    std_all = [0.03, 0.04, 0.05, 0.06]
    std = std_all[i-1]
    input_files = ['data/test_objectdataset_augmentedrot_scale75.h5']
    output_folder = 'data/h5_files/main_split'
    for input_file in input_files:
        # add_noise_and_save(input_file, output_folder)
        with h5py.File(input_file, 'r') as f:
            original_data = f['data'][:]
            labels = f['label'][:]
        np.random.seed(random_seed)
        gaussian_noisy_data = original_data + np.random.normal(mean, std, original_data.shape)
        output_file = os.path.join(output_folder, os.path.basename(input_file))
        with h5py.File(output_file, 'w') as f:
            f.create_dataset('data', data=gaussian_noisy_data)
            f.create_dataset('label', data=labels)

def dropout(x, random_seed):
    c = [(10, 70), (10, 80), (10, 90), (10, 100)]
    # c = [(4, 100), (6, 100), (8, 100), (10, 100)]
    drop_point = c[x-1]
    input_files = ['data/test_objectdataset_augmentedrot_scale75.h5']
    output_folder = 'data/h5_files/main_split'
    for input_file in input_files:
        with h5py.File(input_file, 'r') as f:
            original_data = f['data'][:]  # Assuming 'data' is the dataset name, update accordingly
            labels = f['label'][:]  # Assuming 'label' is the dataset name, update accordingly
        dropout_data = []
        np.random.seed(random_seed)
        for i in range(original_data.shape[0]):
            pointcloud = original_data[i]
            for _ in range(drop_point[0]):

                a = np.random.choice(pointcloud.shape[0], 1)
                picked = pointcloud[a]
                dist = np.sum((pointcloud-picked) ** 2, axis=1, keepdims=True)
                idx = np.argpartition(dist, drop_point[1], axis=0)[:drop_point[1]]

                pointcloud = np.delete(pointcloud, idx.squeeze(), axis=0)

            dropout_data.append(pointcloud)
        output_file = os.path.join(output_folder, os.path.basename(input_file))
        with h5py.File(output_file, 'w') as f:
            f.create_dataset('data', data=dropout_data)
            f.create_dataset('label', data=labels)

# def rotation(i, random_seed):
#     all = [10, 8, 7, 6]
#     p = all[i-1]
#
#     input_files = ['data/test_objectdataset_augmentedrot_scale75.h5']
#     output_folder = 'data/h5_files/main_split'
#     for input_file in input_files:
#         with h5py.File(input_file, 'r') as f:
#             original_data = f['data'][:]  # Assuming 'data' is the dataset name, update accordingly
#             labels = f['label'][:]
#
#         transform_1 = np.identity(4, dtype=np.float32)
#         theta = np.pi / p
#
#         transform_1[0, 0] = np.cos(theta)
#         transform_1[0, 1] = -np.sin(theta)
#         transform_1[1, 0] = np.sin(theta)
#         transform_1[1, 1] = np.cos(theta)
#
#         transformed_data = np.empty_like(original_data)
#
#         for i in range(original_data.shape[0]):
#             point_cloud = original_data[i, :, :]
#
#             transformed_point_cloud = np.dot(point_cloud, transform_1[:3, :3].T) + transform_1[:3, 3]
#
#             transformed_data[i, :, :] = transformed_point_cloud
#         rotated_data = transformed_data
#
#         output_file = os.path.join(output_folder, os.path.basename(input_file))
#         with h5py.File(output_file, 'w') as f:
#             f.create_dataset('data', data=rotated_data)
#             f.create_dataset('label', data=labels)



def rotation(i, random_seed):
    all_angle = [20, 25, 30, 35]
    rotation_angle = all_angle[i-1]
    input_files = ['data/test_objectdataset_augmentedrot_scale75.h5']
    output_folder = 'data/h5_files/main_split'

    for input_file in input_files:
        with h5py.File(input_file, 'r') as f:
            original_data = f['data'][:]
            labels = f['label'][:]
        data = []
        for i in range(original_data.shape[0]):
            pointcloud = original_data[i]
            # rotation_angle = np.radians(30)  # Rotation angle in radians
            # random_seed = np.random.randint(1, 10000)  # Replace with your desired random seed

            rotated_point_cloud = add_random_rotation_noise(pointcloud, np.radians(rotation_angle), random_seed)
            data.append(rotated_point_cloud)
        output_file = os.path.join(output_folder, os.path.basename(input_file))
        with h5py.File(output_file, 'w') as f:
            f.create_dataset('data', data=data)
            f.create_dataset('label', data=labels)

def add_random_rotation_noise(point_cloud, rotation_angle, random_seed=None):
    """
    Add random rotation noise to a point cloud.

    Parameters:
    - point_cloud (numpy array): Input point cloud of shape (N, 3) where N is the number of points.
    - rotation_angle (float): Rotation angle in radians.
    - random_seed (int): Random seed for reproducibility.

    Returns:
    - rotated_point_cloud (numpy array): Point cloud after applying random rotation.
    """
    if random_seed is not None:
        np.random.seed(random_seed)

    # Generate a random rotation axis using random seed
    rotation_axis = np.random.rand(3)
    rotation_axis /= np.linalg.norm(rotation_axis)

    # Generate a random rotation matrix using axis-angle representation
    rotation_matrix = rotation_matrix_from_axis_angle(rotation_axis, rotation_angle)

    # Apply rotation to the point cloud
    rotated_point_cloud = np.dot(point_cloud, rotation_matrix.T)

    return rotated_point_cloud

def rotation_matrix_from_axis_angle(axis, angle):
    """
    Generate a 3x3 rotation matrix from an axis and an angle using Rodrigues' formula.

    Parameters:
    - axis (numpy array): Rotation axis of shape (3,).
    - angle (float): Rotation angle in radians.

    Returns:
    - rotation_matrix (numpy array): 3x3 rotation matrix.
    """
    axis /= np.linalg.norm(axis)
    cross_product_matrix = np.array([[0, -axis[2], axis[1]],
                                     [axis[2], 0, -axis[0]],
                                     [-axis[1], axis[0], 0]])
    rotation_matrix = np.identity(3) + np.sin(angle) * cross_product_matrix + (1 - np.cos(angle)) * np.dot(cross_product_matrix, cross_product_matrix)
    return rotation_matrix



if __name__ == '__main__':
    main()
